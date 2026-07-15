from __future__ import annotations

import datetime as dt
import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .atomic_files import atomic_write_json
from .db import fetch_products
from .filler import (
    class81_index,
    find_article_index,
    is_reportable_header,
    is_yellow_header_cell,
    normalize_article,
    normalize_header,
    product_value,
    product_with_template_class,
)


@dataclass(frozen=True)
class ManualImportRow:
    article: str
    field: str
    value: str
    previous_value: str
    action: str


@dataclass(frozen=True)
class ManualImportResult:
    filled_file: Path
    values_path: Path
    report_path: Path
    imported_values: int
    skipped_values: int


def manual_values_path(work_dir: Path) -> Path:
    return work_dir / "rules" / "manual_values.json"


def normalize_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return " ".join(text.split())


def read_manual_document(path: Path) -> dict:
    if not path.exists():
        return {"version": 1, "values": []}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("manual_values.json должен содержать JSON-объект.")
    data.setdefault("version", 1)
    data.setdefault("values", [])
    if not isinstance(data["values"], list):
        raise ValueError("manual_values.json поле values должно быть списком.")
    return data


def write_manual_document(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    atomic_write_json(path, data)


def load_manual_values(path: Path | None) -> dict[tuple[str, str], str]:
    if path is None:
        return {}
    data = read_manual_document(path)
    values: dict[tuple[str, str], str] = {}
    for item in data.get("values", []):
        if not isinstance(item, dict):
            continue
        article = normalize_value(item.get("article"))
        field = normalize_value(item.get("field"))
        value = normalize_value(item.get("value"))
        if article and field and value:
            values[(article, field)] = value
    return values


def upsert_manual_values(path: Path, rows: list[ManualImportRow], source_file: Path) -> None:
    data = read_manual_document(path)
    imported_at = dt.datetime.now().isoformat(timespec="seconds")
    by_key: dict[tuple[str, str], dict] = {}
    for item in data["values"]:
        if not isinstance(item, dict):
            continue
        key = (normalize_value(item.get("article")), normalize_value(item.get("field")))
        if key[0] and key[1]:
            by_key[key] = item

    for row in rows:
        by_key[(row.article, row.field)] = {
            "article": row.article,
            "field": row.field,
            "value": row.value,
            "source_file": str(source_file),
            "imported_at": imported_at,
        }

    data["values"] = sorted(by_key.values(), key=lambda item: (item["article"], item["field"]))
    write_manual_document(path, data)


def write_manual_import_report(path: Path, rows: list[ManualImportRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Импортировано"
    sheet.append(["Артикул", "Поле", "Ручное значение", "Было бы из базы", "Действие"])
    for row in rows:
        sheet.append([row.article, row.field, row.value, row.previous_value, row.action])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 70)
    workbook.save(path)


def import_manual_completions(
    db_path: Path,
    filled_path: Path,
    values_path: Path,
    output_dir: Path,
    rules_path: Path | None = None,
) -> ManualImportResult:
    if filled_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Импорт ручных дозаполнений пока поддерживает только XLSX/XLSM.")
    if not filled_path.exists():
        raise FileNotFoundError(f"Файл не найден: {filled_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"База не найдена: {db_path}")

    workbook = load_workbook(filled_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_cells = list(sheet[1])
    headers = [normalize_header(cell.value) for cell in header_cells]
    fillable_columns = {
        idx
        for idx, cell in enumerate(header_cells)
        if is_yellow_header_cell(cell)
    }
    article_idx = find_article_index(headers)
    if article_idx is None:
        workbook.close()
        raise ValueError("Не найдена колонка Артикул или Расширенный артикул.")
    class_idx = class81_index(headers)

    articles: list[str] = []
    rows = list(sheet.iter_rows(min_row=2))
    for row in rows:
        if article_idx < len(row):
            articles.append(normalize_article(row[article_idx].value))
    products = fetch_products(db_path, articles)

    imported: list[ManualImportRow] = []
    skipped = 0
    for row in rows:
        article = normalize_article(row[article_idx].value if article_idx < len(row) else "")
        if not article:
            continue
        product = products.get(article)
        if product and class_idx is not None and class_idx < len(row):
            product = product_with_template_class(product, normalize_header(row[class_idx].value))
        for col_idx, header in enumerate(headers):
            if col_idx >= len(row) or col_idx == article_idx:
                continue
            if col_idx not in fillable_columns or not is_reportable_header(header):
                continue
            manual_value = normalize_value(row[col_idx].value)
            if not manual_value:
                continue
            db_value = ""
            if product:
                db_value, _, _ = product_value(product, header, rules_path)
                db_value = normalize_value(db_value)
            if db_value == manual_value:
                skipped += 1
                continue
            imported.append(
                ManualImportRow(
                    article=article,
                    field=header,
                    value=manual_value,
                    previous_value=db_value,
                    action="Сохранено как ручное дозаполнение",
                )
            )
    workbook.close()

    upsert_manual_values(values_path, imported, filled_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"{filled_path.stem}_manual_import_{dt.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    write_manual_import_report(report_path, imported)
    return ManualImportResult(
        filled_file=filled_path,
        values_path=values_path,
        report_path=report_path,
        imported_values=len(imported),
        skipped_values=skipped,
    )


def result_lines(result: ManualImportResult) -> list[str]:
    return [
        f"Файл: {result.filled_file}",
        f"Ручных значений сохранено: {result.imported_values}",
        f"Пропущено без изменений: {result.skipped_values}",
        f"Файл ручных значений: {result.values_path}",
        f"Отчет: {result.report_path}",
    ]
