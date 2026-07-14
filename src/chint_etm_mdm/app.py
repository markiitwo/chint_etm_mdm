from __future__ import annotations

import re
import shutil
import sys
import traceback
from collections import Counter, defaultdict
from pathlib import Path

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from openpyxl import load_workbook

from .analyzer import analyze_template_mapping
from .config import AppConfig, ensure_work_dirs, load_config, save_config
from .db import get_stats
from .filler import FillResult, fill_template
from .mapping_rules import (
    add_approved_class_rule,
    add_rejected_class_rule,
    ensure_default_rules,
    workdir_rules_path,
)


CATEGORY_FROM_FILENAME_RE = re.compile(r"(?:category_template|template)_(\d{4,})", re.IGNORECASE)


class FillWorker(QThread):
    done = Signal(object)
    failed = Signal(str)

    def __init__(
        self, db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None
    ) -> None:
        super().__init__()
        self.db_path = db_path
        self.template_path = template_path
        self.output_dir = output_dir
        self.rules_path = rules_path

    def run(self) -> None:
        try:
            self.done.emit(
                fill_template(self.db_path, self.template_path, self.output_dir, self.rules_path)
            )
        except Exception:
            self.failed.emit(traceback.format_exc())


class AnalyzeWorker(QThread):
    done = Signal(object)
    failed = Signal(str)

    def __init__(
        self, db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None
    ) -> None:
        super().__init__()
        self.db_path = db_path
        self.template_path = template_path
        self.output_dir = output_dir
        self.rules_path = rules_path

    def run(self) -> None:
        try:
            self.done.emit(
                analyze_template_mapping(
                    self.db_path, self.template_path, self.output_dir, self.rules_path
                )
            )
        except Exception:
            self.failed.emit(traceback.format_exc())


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("CHINT ETM MDM")
        self.resize(980, 680)
        self.config = load_config()
        self.worker: FillWorker | AnalyzeWorker | None = None
        self.tabs: QTabWidget | None = None

        self.work_dir_edit = QLineEdit(self.config.work_dir)
        self.db_path_edit = QLineEdit(self.config.db_path)
        self.template_path_edit = QLineEdit()
        self.output_dir_edit = QLineEdit(self.default_output_dir())
        self.mapping_review_path_edit = QLineEdit()
        self.coverage_table = QTableWidget(0, 8)
        self.rules_table = QTableWidget(0, 8)
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.fill_log = QTextEdit()
        self.fill_log.setReadOnly(True)
        self.rules_log = QTextEdit()
        self.rules_log.setReadOnly(True)

        self.setCentralWidget(self.build_ui())
        self.refresh_database_status()

    def build_ui(self) -> QWidget:
        self.tabs = QTabWidget()
        self.tabs.addTab(self.build_database_tab(), "База")
        self.tabs.addTab(self.build_fill_tab(), "Заполнение upload_goods")
        self.tabs.addTab(self.build_rules_tab(), "Правила маппинга")

        root = QWidget()
        layout = QVBoxLayout(root)
        layout.addWidget(self.tabs)
        return root

    def build_database_tab(self) -> QWidget:
        root = QWidget()
        layout = QVBoxLayout(root)

        setup_box = QGroupBox("Рабочая папка и база")
        form = QGridLayout(setup_box)
        form.addWidget(QLabel("Рабочая папка"), 0, 0)
        form.addWidget(self.work_dir_edit, 0, 1)
        choose_work = QPushButton("Выбрать...")
        choose_work.clicked.connect(self.choose_work_dir)
        form.addWidget(choose_work, 0, 2)

        form.addWidget(QLabel("База SQLite"), 1, 0)
        form.addWidget(self.db_path_edit, 1, 1)
        choose_db = QPushButton("Выбрать...")
        choose_db.clicked.connect(self.choose_db)
        form.addWidget(choose_db, 1, 2)

        buttons = QHBoxLayout()
        save_button = QPushButton("Сохранить настройки")
        save_button.clicked.connect(self.save_current_config)
        backup_button = QPushButton("Копировать базу в рабочую папку")
        backup_button.clicked.connect(self.copy_db_to_workspace)
        refresh_button = QPushButton("Обновить статус")
        refresh_button.clicked.connect(self.refresh_database_status)
        buttons.addWidget(save_button)
        buttons.addWidget(backup_button)
        buttons.addWidget(refresh_button)

        layout.addWidget(setup_box)
        layout.addLayout(buttons)
        layout.addWidget(QLabel("Статус базы"))
        layout.addWidget(self.status_text, stretch=1)
        return root

    def build_fill_tab(self) -> QWidget:
        root = QWidget()
        layout = QVBoxLayout(root)

        box = QGroupBox("Шаблон")
        form = QFormLayout(box)

        template_row = QHBoxLayout()
        template_row.addWidget(self.template_path_edit)
        choose_template = QPushButton("Выбрать...")
        choose_template.clicked.connect(self.choose_template)
        template_row.addWidget(choose_template)
        form.addRow("Файл upload_goods", template_row)

        output_row = QHBoxLayout()
        output_row.addWidget(self.output_dir_edit)
        choose_output = QPushButton("Выбрать...")
        choose_output.clicked.connect(self.choose_output_dir)
        output_row.addWidget(choose_output)
        form.addRow("Папка результата", output_row)

        actions = QHBoxLayout()
        run_button = QPushButton("Заполнить из базы")
        run_button.clicked.connect(self.run_fill)
        analyze_button = QPushButton("Проанализировать маппинг")
        analyze_button.clicked.connect(self.run_mapping_analysis)
        actions.addWidget(run_button)
        actions.addWidget(analyze_button)
        actions.addStretch(1)

        layout.addWidget(box)
        layout.addLayout(actions)
        layout.addWidget(QLabel("Журнал"))
        layout.addWidget(self.fill_log, stretch=1)
        return root

    def build_rules_tab(self) -> QWidget:
        root = QWidget()
        layout = QVBoxLayout(root)

        box = QGroupBox("Отчет mapping_review")
        form = QFormLayout(box)
        review_row = QHBoxLayout()
        review_row.addWidget(self.mapping_review_path_edit)
        choose_review = QPushButton("Выбрать...")
        choose_review.clicked.connect(self.choose_mapping_review)
        review_row.addWidget(choose_review)
        load_review = QPushButton("Загрузить кандидаты")
        load_review.clicked.connect(self.load_mapping_review)
        review_row.addWidget(load_review)
        form.addRow("Файл отчета", review_row)

        self.coverage_table.setHorizontalHeaderLabels(
            [
                "Статус",
                "Категория",
                "Поле шаблона",
                "Товаров",
                "Заполнится",
                "Нужен выбор источника",
                "К продактам",
                "Комментарий",
            ]
        )
        self.coverage_table.setColumnWidth(0, 150)
        self.coverage_table.setColumnWidth(1, 90)
        self.coverage_table.setColumnWidth(2, 230)
        self.coverage_table.setColumnWidth(3, 70)
        self.coverage_table.setColumnWidth(4, 80)
        self.coverage_table.setColumnWidth(5, 105)
        self.coverage_table.setColumnWidth(6, 90)
        self.coverage_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)

        self.rules_table.setHorizontalHeaderLabels(
            [
                "Добавить",
                "81 класс",
                "Поле шаблона",
                "Источник",
                "Примеры значений",
                "Товаров",
                "Принять",
                "Отклонить",
            ]
        )
        self.rules_table.setColumnWidth(0, 80)
        self.rules_table.setColumnWidth(1, 100)
        self.rules_table.setColumnWidth(2, 240)
        self.rules_table.setColumnWidth(3, 220)
        self.rules_table.setColumnWidth(4, 300)
        self.rules_table.setColumnWidth(5, 80)
        self.rules_table.setColumnWidth(6, 100)
        self.rules_table.setColumnWidth(7, 110)
        self.rules_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)

        actions = QHBoxLayout()
        save_rules = QPushButton("Сохранить выбранные правила")
        save_rules.clicked.connect(self.save_selected_rules)
        reject_rules = QPushButton("Отклонить выбранные")
        reject_rules.clicked.connect(self.reject_selected_rules)
        actions.addWidget(save_rules)
        actions.addWidget(reject_rules)
        actions.addStretch(1)

        layout.addWidget(box)
        layout.addWidget(QLabel("Покрытие желтых полей"))
        layout.addWidget(self.coverage_table, stretch=1)
        layout.addWidget(QLabel("Кандидаты на выбор источника"))
        layout.addWidget(self.rules_table, stretch=1)
        layout.addLayout(actions)
        layout.addWidget(QLabel("Журнал правил"))
        layout.addWidget(self.rules_log, stretch=1)
        return root

    def default_output_dir(self) -> str:
        if self.config.work_dir:
            return str(Path(self.config.work_dir) / "output")
        return ""

    def choose_work_dir(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "Выберите рабочую папку")
        if not path:
            return
        work_dir = Path(path)
        ensure_work_dirs(work_dir)
        self.work_dir_edit.setText(str(work_dir))
        self.output_dir_edit.setText(str(work_dir / "output"))
        self.save_current_config()

    def choose_db(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите chint_mdm.sqlite",
            "",
            "SQLite database (*.sqlite *.sqlite3 *.db);;All files (*)",
        )
        if path:
            self.db_path_edit.setText(path)
            self.save_current_config()
            self.refresh_database_status()

    def choose_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите upload_goods шаблон",
            "",
            "ETM templates (*.xlsx *.xlsm *.csv);;All files (*)",
        )
        if path:
            self.template_path_edit.setText(path)

    def choose_output_dir(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "Выберите папку результата")
        if path:
            self.output_dir_edit.setText(path)

    def choose_mapping_review(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите mapping_review.xlsx",
            "",
            "Mapping review (*.xlsx *.xlsm);;All files (*)",
        )
        if path:
            self.mapping_review_path_edit.setText(path)
            self.load_mapping_review()

    def save_current_config(self) -> None:
        self.config = AppConfig(
            work_dir=self.work_dir_edit.text().strip(),
            db_path=self.db_path_edit.text().strip(),
        )
        if self.config.work_path:
            ensure_work_dirs(self.config.work_path)
            ensure_default_rules(self.config.work_path)
        save_config(self.config)

    def copy_db_to_workspace(self) -> None:
        source = Path(self.db_path_edit.text().strip())
        work_dir_text = self.work_dir_edit.text().strip()
        if not source.exists():
            QMessageBox.warning(self, "База не найдена", "Сначала выберите существующую базу.")
            return
        if not work_dir_text:
            QMessageBox.warning(self, "Нет рабочей папки", "Сначала выберите рабочую папку.")
            return
        target_dir = Path(work_dir_text) / "database"
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / "chint_mdm.sqlite"
        shutil.copy2(source, target)
        self.db_path_edit.setText(str(target))
        self.save_current_config()
        self.refresh_database_status()
        QMessageBox.information(self, "Готово", f"База скопирована:\n{target}")

    def refresh_database_status(self) -> None:
        db_text = self.db_path_edit.text().strip()
        if not db_text:
            self.status_text.setPlainText("База не выбрана.")
            return
        db_path = Path(db_text)
        if not db_path.exists():
            self.status_text.setPlainText(f"База не найдена:\n{db_path}")
            return
        try:
            stats = get_stats(db_path)
        except Exception as exc:
            self.status_text.setPlainText(f"Не удалось прочитать базу:\n{exc}")
            return
        rules_text = "не выбран"
        work_text = self.work_dir_edit.text().strip()
        if work_text:
            rules_text = str(workdir_rules_path(Path(work_text)))

        self.status_text.setPlainText(
            "\n".join(
                [
                    f"Файл: {db_path}",
                    f"Файл правил: {rules_text}",
                    f"Товаров: {stats.products_count}",
                    f"Габариты/вес: {stats.dimensions_count}",
                    f"ETIM-значений: {stats.attributes_count}",
                    f"Последний прайс: {stats.latest_price_snapshot}",
                    f"Файл прайса: {stats.latest_price_file}",
                ]
            )
        )

    def run_fill(self) -> None:
        try:
            self.save_current_config()
            paths = self.validated_paths()
            if paths is None:
                return
            db_path, template_path, output_dir, rules_path = paths

            self.fill_log.append("Запуск заполнения...")
            self.worker = FillWorker(db_path, template_path, output_dir, rules_path)
            self.worker.done.connect(self.on_fill_done)
            self.worker.failed.connect(self.on_fill_failed)
            self.worker.start()
        except Exception:
            self.on_action_failed(traceback.format_exc())

    def run_mapping_analysis(self) -> None:
        try:
            self.save_current_config()
            paths = self.validated_paths()
            if paths is None:
                return
            db_path, template_path, output_dir, rules_path = paths

            self.fill_log.append("Анализирую желтые поля и кандидаты маппинга...")
            self.worker = AnalyzeWorker(db_path, template_path, output_dir, rules_path)
            self.worker.done.connect(self.on_mapping_analysis_done)
            self.worker.failed.connect(self.on_action_failed)
            self.worker.start()
        except Exception:
            self.on_action_failed(traceback.format_exc())

    def validated_paths(self) -> tuple[Path, Path, Path, Path | None] | None:
        db_path = Path(self.db_path_edit.text().strip())
        template_path = Path(self.template_path_edit.text().strip())
        output_text = self.output_dir_edit.text().strip()
        output_dir = Path(output_text) if output_text else Path()
        work_text = self.work_dir_edit.text().strip()
        rules_path = None
        if work_text:
            rules_path = ensure_default_rules(Path(work_text))

        if not db_path.exists():
            QMessageBox.warning(self, "База не найдена", "Выберите существующую SQLite базу.")
            return None
        if not template_path.exists():
            QMessageBox.warning(self, "Шаблон не найден", "Выберите файл шаблона.")
            return None
        if not output_text:
            QMessageBox.warning(self, "Нет папки результата", "Выберите папку результата.")
            return None
        return db_path, template_path, output_dir, rules_path

    def on_fill_done(self, result: FillResult) -> None:
        self.fill_log.append("Готово.")
        self.fill_log.append(f"Строк в шаблоне: {result.total_rows}")
        self.fill_log.append(f"Найдено артикулов: {result.found_articles}")
        self.fill_log.append(f"Не найдено артикулов: {result.missing_articles}")
        self.fill_log.append(f"Заполнено ячеек: {result.filled_cells}")
        self.fill_log.append(f"Предложений в отчете: {result.suggested_cells}")
        self.fill_log.append(f"Файл: {result.output_path}")
        self.fill_log.append(f"Отчет: {result.report_path}")
        QMessageBox.information(
            self,
            "Заполнение завершено",
            f"Файл:\n{result.output_path}\n\nОтчет:\n{result.report_path}",
        )

    def on_mapping_analysis_done(self, report_path: Path) -> None:
        self.fill_log.append("Анализ маппинга готов.")
        self.fill_log.append(f"Отчет: {report_path}")
        self.mapping_review_path_edit.setText(str(report_path))
        self.load_mapping_review()
        if self.tabs is not None:
            self.tabs.setCurrentIndex(2)
        QMessageBox.information(self, "Анализ готов", f"Отчет:\n{report_path}")

    def on_fill_failed(self, details: str) -> None:
        self.fill_log.append("Ошибка заполнения.")
        self.fill_log.append(details)
        QMessageBox.critical(self, "Ошибка", details)

    def on_action_failed(self, details: str) -> None:
        self.fill_log.append("Ошибка.")
        self.fill_log.append(details)
        QMessageBox.critical(self, "Ошибка", details)

    def load_mapping_review(self) -> None:
        path_text = self.mapping_review_path_edit.text().strip()
        if not path_text:
            QMessageBox.warning(self, "Отчет не выбран", "Выберите файл mapping_review.xlsx.")
            return
        review_path = Path(path_text)
        if not review_path.exists():
            QMessageBox.warning(self, "Отчет не найден", "Выберите существующий файл mapping_review.xlsx.")
            return

        try:
            coverage_rows = self.read_mapping_review_coverage(review_path)
            rows = self.read_mapping_review_candidates(review_path)
        except Exception as exc:
            QMessageBox.critical(self, "Ошибка чтения отчета", str(exc))
            return

        self.coverage_table.setRowCount(0)
        for row_data in coverage_rows:
            row_idx = self.coverage_table.rowCount()
            self.coverage_table.insertRow(row_idx)
            values = [
                row_data["status"],
                row_data["class81_code"],
                row_data["template_field"],
                row_data["products_count"],
                row_data["will_fill"],
                row_data["needs_source_choice"],
                row_data["needs_pm"],
                row_data["note"],
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.coverage_table.setItem(row_idx, col_idx, item)

        self.rules_table.setRowCount(0)
        for row_data in rows:
            row_idx = self.rules_table.rowCount()
            self.rules_table.insertRow(row_idx)

            check_item = QTableWidgetItem()
            check_item.setFlags(
                Qt.ItemFlag.ItemIsUserCheckable
                | Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsSelectable
            )
            check_item.setCheckState(Qt.CheckState.Unchecked)
            check_item.setData(Qt.ItemDataRole.UserRole, row_data)
            self.rules_table.setItem(row_idx, 0, check_item)

            values = [
                row_data["class81_code"],
                row_data["template_field"],
                row_data["source_attribute"],
                row_data["examples"],
                row_data["coverage"],
            ]
            for col_idx, value in enumerate(values, start=1):
                item = QTableWidgetItem(str(value))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.rules_table.setItem(row_idx, col_idx, item)

            accept_button = QPushButton("Принять")
            accept_button.clicked.connect(
                lambda _checked=False, data=dict(row_data): self.save_rule_data(data)
            )
            self.rules_table.setCellWidget(row_idx, 6, accept_button)

            reject_button = QPushButton("Отклонить")
            reject_button.clicked.connect(
                lambda _checked=False, data=dict(row_data): self.reject_rule_data(data)
            )
            self.rules_table.setCellWidget(row_idx, 7, reject_button)

        pm_count = sum(int(row["needs_pm"] or "0") for row in coverage_rows if str(row["needs_pm"]).isdigit())
        mapping_count = sum(
            int(row["needs_source_choice"] or "0")
            for row in coverage_rows
            if str(row["needs_source_choice"]).isdigit()
        )
        self.rules_log.append(
            f"Загружено полей: {len(coverage_rows)}; кандидатов: {len(rows)}; "
            f"к продактам строк: {pm_count}; нужен выбор источника: {mapping_count}"
        )
        if mapping_count and not rows:
            self.rules_log.append(
                "Есть строки, где нужен выбор источника, но кандидаты не найдены. "
                "Скорее всего, в отчете не указана категория товара. "
                "Пересоздайте анализ маппинга на новой версии программы."
            )

    def infer_category_from_review_path(self, review_path: Path) -> str:
        match = CATEGORY_FROM_FILENAME_RE.search(review_path.name)
        return match.group(1) if match else ""

    def read_mapping_review_coverage(self, review_path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(review_path, data_only=True, read_only=True)
        if "Покрытие" not in workbook.sheetnames:
            raise ValueError("В отчете нет листа 'Покрытие'. Сначала сделайте анализ маппинга.")
        sheet = workbook["Покрытие"]
        header_row = next(sheet.iter_rows(values_only=True), None)
        if not header_row:
            raise ValueError("Лист 'Покрытие' пустой.")
        headers = [str(value or "").strip() for value in header_row]
        index = {name: idx for idx, name in enumerate(headers)}
        category_column = "Категория" if "Категория" in index else "81 класс"
        required = [
            category_column,
            "Поле шаблона",
            "Статус",
            "Товаров",
            "Заполнится",
            "Нужен выбор источника",
            "К продактам",
            "Комментарий",
        ]
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"В листе 'Покрытие' нет колонок: {', '.join(missing)}")

        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            class81_code = str(values[index[category_column]] or "").strip()
            if not class81_code:
                class81_code = self.infer_category_from_review_path(review_path)
            template_field = str(values[index["Поле шаблона"]] or "").strip()
            if not class81_code and not template_field:
                continue
            rows.append(
                {
                    "class81_code": class81_code,
                    "template_field": template_field,
                    "status": str(values[index["Статус"]] or "").strip(),
                    "products_count": str(values[index["Товаров"]] or "").strip(),
                    "will_fill": str(values[index["Заполнится"]] or "").strip(),
                    "needs_source_choice": str(
                        values[index["Нужен выбор источника"]] or ""
                    ).strip(),
                    "needs_pm": str(values[index["К продактам"]] or "").strip(),
                    "note": str(values[index["Комментарий"]] or "").strip(),
                }
            )
        return rows

    def read_mapping_review_candidates(self, review_path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(review_path, data_only=True, read_only=True)
        if "Правила" not in workbook.sheetnames:
            raise ValueError("В отчете нет листа 'Правила'. Сначала сделайте анализ маппинга.")
        sheet = workbook["Правила"]
        header_row = next(sheet.iter_rows(values_only=True), None)
        if not header_row:
            raise ValueError("Лист 'Правила' пустой.")
        headers = [str(value or "").strip() for value in header_row]
        index = {name: idx for idx, name in enumerate(headers)}
        required = ["81 класс", "Поле шаблона", "Кандидат источника", "Покрытие кандидата", "Действие"]
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"В листе 'Правила' нет колонок: {', '.join(missing)}")

        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            action = str(values[index["Действие"]] or "").strip()
            if "уже утверждено" in action.lower():
                continue
            class81_code = str(values[index["81 класс"]] or "").strip()
            if not class81_code:
                class81_code = self.infer_category_from_review_path(review_path)
            template_field = str(values[index["Поле шаблона"]] or "").strip()
            source_attribute = str(values[index["Кандидат источника"]] or "").strip()
            coverage = str(values[index["Покрытие кандидата"]] or "").strip()
            examples = ""
            if "Примеры значений" in index:
                examples = str(values[index["Примеры значений"]] or "").strip()
            if not class81_code or not template_field or not source_attribute:
                continue
            rows.append(
                {
                    "class81_code": class81_code,
                    "template_field": template_field,
                    "source_attribute": source_attribute,
                    "coverage": coverage,
                    "examples": examples,
                    "action": action,
                }
            )
        if rows:
            return rows
        return self.read_mapping_review_candidates_from_source_choice(workbook, review_path)

    def read_mapping_review_candidates_from_source_choice(
        self, workbook, review_path: Path
    ) -> list[dict[str, str]]:
        if "Выбор источника" not in workbook.sheetnames:
            return []
        sheet = workbook["Выбор источника"]
        header_row = next(sheet.iter_rows(values_only=True), None)
        if not header_row:
            return []
        headers = [str(value or "").strip() for value in header_row]
        index = {name: idx for idx, name in enumerate(headers)}
        category_column = "Категория" if "Категория" in index else "81 класс"
        required = [category_column, "Поле шаблона", "Возможный источник", "Значение", "Артикул"]
        if any(name not in index for name in required):
            return []

        counts: Counter[tuple[str, str, str]] = Counter()
        examples: dict[tuple[str, str, str], list[str]] = defaultdict(list)
        articles: dict[tuple[str, str, str], list[str]] = defaultdict(list)
        for values in sheet.iter_rows(min_row=2, values_only=True):
            class81_code = str(values[index[category_column]] or "").strip()
            if not class81_code:
                class81_code = self.infer_category_from_review_path(review_path)
            template_field = str(values[index["Поле шаблона"]] or "").strip()
            source_attribute = str(values[index["Возможный источник"]] or "").strip()
            value = str(values[index["Значение"]] or "").strip()
            article = str(values[index["Артикул"]] or "").strip()
            if not class81_code or not template_field or not source_attribute:
                continue
            key = (class81_code, template_field, source_attribute)
            counts[key] += 1
            if value and value not in examples[key]:
                examples[key].append(value)
            if article and article not in articles[key]:
                articles[key].append(article)

        rows: list[dict[str, str]] = []
        for (class81_code, template_field, source_attribute), count in counts.most_common():
            rows.append(
                {
                    "class81_code": class81_code,
                    "template_field": template_field,
                    "source_attribute": source_attribute,
                    "coverage": str(count),
                    "examples": "; ".join(examples[(class81_code, template_field, source_attribute)][:5]),
                    "action": "собрано из листа Выбор источника",
                }
            )
        return rows

    def rules_path_from_work_dir(self) -> Path | None:
        work_text = self.work_dir_edit.text().strip()
        if not work_text:
            QMessageBox.warning(self, "Нет рабочей папки", "Сначала выберите рабочую папку на вкладке База.")
            return None
        return ensure_default_rules(Path(work_text))

    def row_data(self, row_idx: int) -> dict[str, str] | None:
        check_item = self.rules_table.item(row_idx, 0)
        if not check_item:
            return None
        row_data = check_item.data(Qt.ItemDataRole.UserRole)
        return row_data if isinstance(row_data, dict) else None

    def remove_matching_rule_row(self, target: dict[str, str]) -> None:
        for row_idx in range(self.rules_table.rowCount()):
            row_data = self.row_data(row_idx)
            if row_data is target or row_data == target:
                self.rules_table.removeRow(row_idx)
                return

    def save_rule_data(self, row_data: dict[str, str]) -> bool:
        rules_path = self.rules_path_from_work_dir()
        if rules_path is None:
            return False
        changed = add_approved_class_rule(
            rules_path,
            row_data["class81_code"],
            row_data["template_field"],
            row_data["source_attribute"],
        )
        status = "добавлено" if changed else "уже было"
        self.rules_log.append(
            f"Принято: {row_data['class81_code']} | {row_data['template_field']} <- "
            f"{row_data['source_attribute']} ({status})"
        )
        self.remove_matching_rule_row(row_data)
        return changed

    def save_rule_from_row(self, row_idx: int) -> bool:
        row_data = self.row_data(row_idx)
        return self.save_rule_data(row_data) if row_data is not None else False

    def reject_rule_data(self, row_data: dict[str, str]) -> bool:
        rules_path = self.rules_path_from_work_dir()
        if rules_path is None:
            return False
        changed = add_rejected_class_rule(
            rules_path,
            row_data["class81_code"],
            row_data["template_field"],
            row_data["source_attribute"],
        )
        status = "добавлено в игнор" if changed else "уже было в игноре"
        self.rules_log.append(
            f"Отклонено: {row_data['class81_code']} | {row_data['template_field']} <- "
            f"{row_data['source_attribute']} ({status})"
        )
        self.remove_matching_rule_row(row_data)
        return changed

    def reject_rule_from_row(self, row_idx: int) -> bool:
        row_data = self.row_data(row_idx)
        return self.reject_rule_data(row_data) if row_data is not None else False

    def save_selected_rules(self) -> None:
        rules_path = self.rules_path_from_work_dir()
        if rules_path is None:
            return

        added = 0
        skipped = 0
        selected = 0
        selected_rows: list[int] = []
        for row_idx in range(self.rules_table.rowCount()):
            check_item = self.rules_table.item(row_idx, 0)
            if not check_item or check_item.checkState() != Qt.CheckState.Checked:
                continue
            selected += 1
            selected_rows.append(row_idx)
            row_data = check_item.data(Qt.ItemDataRole.UserRole)
            if not isinstance(row_data, dict):
                continue
            changed = add_approved_class_rule(
                rules_path,
                row_data["class81_code"],
                row_data["template_field"],
                row_data["source_attribute"],
            )
            if changed:
                added += 1
            else:
                skipped += 1

        if selected == 0:
            QMessageBox.information(self, "Ничего не выбрано", "Отметьте кандидаты, которые нужно утвердить.")
            return

        self.rules_log.append(f"Файл правил: {rules_path}")
        self.rules_log.append(f"Добавлено правил/источников: {added}; уже было: {skipped}")
        for row_idx in reversed(selected_rows):
            self.rules_table.removeRow(row_idx)
        QMessageBox.information(
            self,
            "Правила сохранены",
            f"Файл правил:\n{rules_path}\n\nДобавлено: {added}\nУже было: {skipped}",
        )

    def reject_selected_rules(self) -> None:
        rules_path = self.rules_path_from_work_dir()
        if rules_path is None:
            return

        rejected = 0
        skipped = 0
        selected_rows: list[int] = []
        for row_idx in range(self.rules_table.rowCount()):
            check_item = self.rules_table.item(row_idx, 0)
            if check_item and check_item.checkState() == Qt.CheckState.Checked:
                selected_rows.append(row_idx)

        if not selected_rows:
            QMessageBox.information(self, "Ничего не выбрано", "Отметьте кандидаты, которые нужно отклонить.")
            return

        for row_idx in reversed(selected_rows):
            row_data = self.row_data(row_idx)
            if not row_data:
                continue
            changed = add_rejected_class_rule(
                rules_path,
                row_data["class81_code"],
                row_data["template_field"],
                row_data["source_attribute"],
            )
            if changed:
                rejected += 1
            else:
                skipped += 1
            self.rules_table.removeRow(row_idx)

        self.rules_log.append(f"Отклонено кандидатов: {rejected}; уже было в игноре: {skipped}")
        QMessageBox.information(
            self,
            "Кандидаты отклонены",
            f"Файл правил:\n{rules_path}\n\nОтклонено: {rejected}\nУже было: {skipped}",
        )


def main() -> None:
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
