from __future__ import annotations

import datetime as dt
import re
import csv
import shutil
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill

from .db import ProductRecord, fetch_products
from .mapping_rules import source_attributes_for


ARTICLE_HEADERS = ("Артикул", "Расширенный артикул")
CONFIDENT_STATIC_VALUES = {
    "Код производителя": "CHINT",
    "Страна": "CHN",
    "Название упаковки": "шт",
}
DIRECT_DB_HEADERS = {
    "Расширенный артикул",
    "81 класс",
    "Название",
    "Полное название",
    "Краткое имя WMS",
    "Вес, кг",
    "Длина, м",
    "Ширина, м",
    "Высота, м",
    "Объем, м3",
    "Код ТН ВЭД",
    "Код ОКПД2",
    "Упак3 Название",
    "Упак3 Емкость",
    "Упак3 Вес, кг",
    "Упак3 Длина, м",
    "Упак3 Ширина, м",
    "Упак3 Высота, м",
    "Упак3 Объем, м3",
    "Штрих-код3",
    "Штрих-код3уп",
    "Отгрузка кратно УпЗавод",
    "Упак5 Емкость",
    "Упак5 Вес, кг",
    "Упак5 Длина, м",
    "Упак5 Ширина, м",
    "Упак5 Высота, м",
    "Упак5 Объем, м3",
    "Штрих-код5",
    "Штрих-код5уп",
}
MISSING_CELL_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
MISSING_CELL_FONT = Font(color="FF9C0006")
DATA_VALIDATIONS_RE = re.compile(br"<dataValidations\b.*?</dataValidations>", re.DOTALL)
EXT_LST_RE = re.compile(br"<extLst\b.*?</extLst>", re.DOTALL)
WORKSHEET_ROOT_RE = re.compile(br"<worksheet\b[^>]*>")
XMLNS_ATTR_RE = re.compile(br'\s+xmlns(?::[A-Za-z_][\w.-]*)?="[^"]*"')
MC_IGNORABLE_RE = re.compile(br'\s+mc:Ignorable="([^"]*)"')


@dataclass(frozen=True)
class FillReportRow:
    row_number: int
    article: str
    column: str
    status: str
    value: str
    source: str
    note: str = ""


@dataclass(frozen=True)
class FillResult:
    output_path: Path
    report_path: Path
    total_rows: int
    found_articles: int
    filled_cells: int
    suggested_cells: int
    missing_articles: int


def normalize_header(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def is_yellow_header_cell(cell: Cell) -> bool:
    fill = cell.fill
    if fill.fill_type != "solid":
        return False
    rgb = str(fill.fgColor.rgb or "").upper()
    return rgb in {"FFFFFF00", "FFFF00"} or rgb.endswith("FFFF00")


def normalize_article(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def comparable_text(value: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", value.lower())


def short_wms_name(value: str, limit: int = 60) -> str:
    clean = " ".join(value.split())
    return clean[:limit].rstrip()


def format_decimal(value: float | int | None, digits: int = 6) -> str:
    if value is None:
        return ""
    text = f"{float(value):.{digits}f}".rstrip("0").rstrip(".")
    return text


def mm_to_m(value: float | int | None) -> str:
    if value is None:
        return ""
    return format_decimal(float(value) / 1000.0, 6)


def format_count(value: float | int | None) -> str:
    if value is None:
        return ""
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return format_decimal(number, 6)


def suggested_attribute(
    product: ProductRecord, attr_name: str, rules_path: Path | None = None
) -> tuple[str, str]:
    attrs = product.attributes or {}
    template_field = f"Конфиг:{attr_name}"
    for alias in source_attributes_for(product.class81_code, template_field, rules_path):
        value = attrs.get(alias, "")
        if value:
            return value, alias

    return "", ""


def product_value(
    product: ProductRecord, header: str, rules_path: Path | None = None
) -> tuple[str, str, str]:
    if header in CONFIDENT_STATIC_VALUES:
        return CONFIDENT_STATIC_VALUES[header], "filled", "static"
    if header == "Расширенный артикул":
        return product.article, "filled", "products.article"
    if header == "81 класс":
        return product.class81_code, "filled", "ipro_goods/product_etm_class_suggestions.class81_code"
    if header == "Название":
        return product.name or product.full_name, "filled", "products/ipro_goods.name"
    if header == "Полное название":
        return product.full_name or product.name, "filled", "ipro_goods.full_name"
    if header == "Краткое имя WMS":
        return short_wms_name(product.name or product.full_name), "filled", "generated"
    if header == "Вес, кг":
        return (
            format_decimal(product.weight_kg, 6),
            "filled",
            "product_dimensions_resolved.weight_kg/price_snapshot_items.gross_weight_unit",
        )
    if header == "Длина, м":
        return mm_to_m(product.length_mm), "filled", "product_dimensions_resolved.length_mm"
    if header == "Ширина, м":
        return mm_to_m(product.width_mm), "filled", "product_dimensions_resolved.width_mm"
    if header == "Высота, м":
        return mm_to_m(product.height_mm), "filled", "product_dimensions_resolved.height_mm"
    if header == "Объем, м3":
        return format_decimal(product.volume_m3, 9), "filled", "product_dimensions_resolved/price_snapshot_items.volume_m3"
    if header == "Код ТН ВЭД":
        return product.tnved_code, "filled", "product_tnved_codes.tnved_code"
    if header == "Код ОКПД2":
        return product.okpd2_code, "filled", "product_okpd2_codes.okpd2"
    if header == "Упак3 Название" and product.min_shipment is not None:
        return "УпЗавод", "filled", "generated"
    if header == "Упак3 Емкость":
        return format_count(product.min_shipment), "filled", "price_snapshot_items.min_shipment"
    if header == "Упак3 Вес, кг":
        return format_decimal(product.package_min_weight_kg, 6), "filled", "price_snapshot_items.gross_weight_min"
    if header == "Упак3 Объем, м3":
        return format_decimal(product.package_min_volume_m3, 9), "filled", "price_snapshot_items.min_volume"
    if header == "Штрих-код3":
        return product.gtin_inner, "filled", "material_data_items.gtin_inner"
    if header == "Штрих-код3уп" and product.gtin_inner:
        return "3", "filled", "static"
    if header == "Отгрузка кратно УпЗавод" and product.min_shipment is not None:
        return "ДА", "filled", "generated"
    if header == "Упак5 Емкость":
        return format_count(product.pack_multiple), "filled", "price_snapshot_items.pack_multiple"
    if header == "Упак5 Вес, кг":
        return format_decimal(product.transport_weight_kg, 6), "filled", "price_snapshot_items/material_data_items.transport_weight"
    if header == "Упак5 Длина, м":
        return mm_to_m(product.transport_length_mm), "filled", "material_data_items.outer_length_mm"
    if header == "Упак5 Ширина, м":
        return mm_to_m(product.transport_width_mm), "filled", "material_data_items.outer_width_mm"
    if header == "Упак5 Высота, м":
        return mm_to_m(product.transport_height_mm), "filled", "material_data_items.outer_height_mm"
    if header == "Упак5 Объем, м3":
        return format_decimal(product.transport_volume_m3, 9), "filled", "price_snapshot_items.transport_volume"
    if header == "Штрих-код5":
        return product.gtin_outer, "filled", "material_data_items.gtin_outer"
    if header == "Штрих-код5уп" and product.gtin_outer:
        return "5", "filled", "static"
    if header.startswith("Конфиг:"):
        attr_name = header.removeprefix("Конфиг:").strip()
        value = (product.attributes or {}).get(attr_name, "")
        if value:
            return value, "filled", "product_attribute_values.attribute_name"
        suggestion, source_name = suggested_attribute(product, attr_name, rules_path)
        if suggestion:
            return suggestion, "filled", f"approved_class_rule:{source_name}"
        return "", "blank", ""
    return "", "blank", ""


def is_reportable_header(header: str) -> bool:
    return header in CONFIDENT_STATIC_VALUES or header in DIRECT_DB_HEADERS or header.startswith("Конфиг:")


def missing_note(header: str) -> str:
    if header in {"Вес, кг", "Длина, м", "Ширина, м", "Высота, м", "Объем, м3"}:
        return "В базе нет значения для единицы товара. Нужно заполнить именно изделие, не упаковку."
    if header == "81 класс":
        return "В базе нет категории товара."
    if header == "Код ТН ВЭД":
        return "В базе нет кода ТН ВЭД."
    if header == "Код ОКПД2":
        return "В базе нет кода ОКПД2."
    if header.startswith("Конфиг:"):
        return "В базе нет подходящего значения для этой характеристики."
    if header.startswith("Упак3") or header in {"Штрих-код3", "Штрих-код3уп", "Отгрузка кратно УпЗавод"}:
        return "В базе нет данных заводской упаковки."
    if header.startswith("Упак5") or header in {"Штрих-код5", "Штрих-код5уп"}:
        return "В базе нет данных транспортной упаковки."
    return "В базе нет значения для этого поля."


def report_status_label(status: str) -> str:
    labels = {
        "filled": "Заполнено",
        "filled_suggested": "Заполнено, нужно проверить",
        "missing_value": "Не заполнено",
        "not_found": "Артикул не найден",
        "ok": "Готово",
        "needs_review": "Нужно проверить",
        "incomplete": "Заполнено не полностью",
    }
    return labels.get(status, status)


def product_manager_action(header: str) -> str:
    if header.startswith("Конфиг:"):
        return f"Заполнить характеристику: {header.removeprefix('Конфиг:').strip()}"
    if header:
        return f"Заполнить поле: {header}"
    return "Проверить артикул в базе"


def source_label(source: str) -> str:
    if not source:
        return ""
    if source == "static":
        return "Значение по умолчанию"
    if source == "generated":
        return "Сформировано программой"
    if source.startswith("approved_class_rule:"):
        return f"Выбранный источник: {source.split(':', 1)[1]}"
    if "attribute" in source:
        return "Характеристика товара в базе"
    if "article" in source:
        return "Артикул из шаблона"
    if "class81" in source:
        return "Категория из базы"
    if "name" in source:
        return "Название из базы"
    if "tnved" in source.lower() or "okpd2" in source.lower():
        return "Код из базы"
    if "weight" in source or "length" in source or "width" in source or "height" in source or "volume" in source:
        return "Вес или габариты из базы"
    if "shipment" in source or "gtin" in source or "transport" in source:
        return "Данные упаковки из базы"
    return "База"


def mark_missing_cell(cell: Cell) -> None:
    cell.fill = MISSING_CELL_FILL
    cell.font = MISSING_CELL_FONT


def preserve_worksheet_dropdowns(template_path: Path, output_path: Path) -> None:
    """Restore worksheet validation XML that openpyxl may simplify or drop."""
    if template_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return

    with ZipFile(template_path, "r") as source_zip, ZipFile(output_path, "r") as output_zip:
        source_names = set(source_zip.namelist())
        output_names = set(output_zip.namelist())
        worksheet_names = sorted(
            name
            for name in source_names.intersection(output_names)
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
        replacements: dict[str, bytes] = {}
        for name in worksheet_names:
            source_xml = source_zip.read(name)
            output_xml = output_zip.read(name)
            source_validations = DATA_VALIDATIONS_RE.search(source_xml)
            source_extensions = EXT_LST_RE.search(source_xml)
            if not source_validations and not source_extensions:
                continue

            updated_xml = DATA_VALIDATIONS_RE.sub(b"", output_xml)
            if source_validations:
                updated_xml = insert_after_xml_tag(
                    updated_xml, b"</sheetData>", source_validations.group(0)
                )

            if source_extensions and b"dataValidations" in source_extensions.group(0):
                updated_xml = EXT_LST_RE.sub(b"", updated_xml)
                updated_xml = updated_xml.replace(
                    b"</worksheet>", source_extensions.group(0) + b"</worksheet>", 1
                )

            updated_xml = copy_worksheet_root_namespaces(source_xml, updated_xml)

            if updated_xml != output_xml:
                replacements[name] = updated_xml

        if not replacements:
            return

        with tempfile.NamedTemporaryFile(delete=False, suffix=output_path.suffix) as tmp:
            temp_path = Path(tmp.name)

        try:
            with ZipFile(temp_path, "w", compression=ZIP_DEFLATED) as new_zip:
                for info in output_zip.infolist():
                    data = replacements.get(info.filename)
                    if data is None:
                        data = output_zip.read(info.filename)
                    new_zip.writestr(info, data)
            shutil.move(str(temp_path), output_path)
        finally:
            if temp_path.exists():
                temp_path.unlink()


def insert_after_xml_tag(xml: bytes, tag: bytes, payload: bytes) -> bytes:
    index = xml.find(tag)
    if index == -1:
        return xml.replace(b"</worksheet>", payload + b"</worksheet>", 1)
    insert_at = index + len(tag)
    return xml[:insert_at] + payload + xml[insert_at:]


def copy_worksheet_root_namespaces(source_xml: bytes, output_xml: bytes) -> bytes:
    source_root = WORKSHEET_ROOT_RE.search(source_xml)
    output_root = WORKSHEET_ROOT_RE.search(output_xml)
    if not source_root or not output_root:
        return output_xml

    source_tag = source_root.group(0)
    output_tag = output_root.group(0)
    updated_tag = output_tag

    existing_namespace_names = {
        declaration.split(b"=", 1)[0].strip()
        for declaration in XMLNS_ATTR_RE.findall(output_tag)
    }
    for declaration in XMLNS_ATTR_RE.findall(source_tag):
        namespace_name = declaration.split(b"=", 1)[0].strip()
        if namespace_name not in existing_namespace_names:
            updated_tag = insert_before_tag_end(updated_tag, declaration)
            existing_namespace_names.add(namespace_name)

    source_ignorable = MC_IGNORABLE_RE.search(source_tag)
    output_ignorable = MC_IGNORABLE_RE.search(updated_tag)
    if source_ignorable and not output_ignorable:
        updated_tag = insert_before_tag_end(updated_tag, source_ignorable.group(0))
    elif source_ignorable and output_ignorable:
        values = dict.fromkeys(
            output_ignorable.group(1).split() + source_ignorable.group(1).split()
        )
        merged = b' mc:Ignorable="' + b" ".join(values) + b'"'
        updated_tag = (
            updated_tag[: output_ignorable.start()]
            + merged
            + updated_tag[output_ignorable.end() :]
        )

    if updated_tag == output_tag:
        return output_xml
    return output_xml[: output_root.start()] + updated_tag + output_xml[output_root.end() :]


def insert_before_tag_end(tag: bytes, attribute: bytes) -> bytes:
    end = tag.rfind(b">")
    if end == -1:
        return tag
    return tag[:end] + attribute + tag[end:]


def find_article_index(headers: list[str]) -> int | None:
    for candidate in ARTICLE_HEADERS:
        if candidate in headers:
            return headers.index(candidate)
    return None


def make_output_paths(template_path: Path, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = output_dir / f"{template_path.stem}_filled_{stamp}{template_path.suffix}"
    report_path = output_dir / f"{template_path.stem}_report_{stamp}.xlsx"
    return output_path, report_path


def write_report(path: Path, rows: Iterable[FillReportRow]) -> None:
    rows = list(rows)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Итог"
    details = workbook.create_sheet("Детали")
    product_manager = workbook.create_sheet("К продактам")

    summary.append(
        [
            "Строка",
            "Артикул",
            "Состояние",
            "Заполнено",
            "Заполнено после проверки",
            "Не заполнено",
            "Что проверить",
        ]
    )
    details.append(["Строка", "Артикул", "Поле", "Состояние", "Значение", "Откуда взяли", "Комментарий"])
    product_manager.append(["Строка", "Артикул", "Поле", "Что сделать", "Комментарий"])

    by_row: dict[tuple[int, str], dict[str, object]] = {}
    for item in rows:
        key = (item.row_number, item.article)
        bucket = by_row.setdefault(
            key,
            {
                "filled": 0,
                "suggested": 0,
                "missing": 0,
                "status": "ok",
                "notes": [],
            },
        )
        if item.status == "filled":
            bucket["filled"] = int(bucket["filled"]) + 1
        elif item.status == "filled_suggested":
            bucket["suggested"] = int(bucket["suggested"]) + 1
            bucket["status"] = "needs_review"
        elif item.status == "missing_value":
            bucket["missing"] = int(bucket["missing"]) + 1
            bucket["status"] = "incomplete"
        elif item.status == "not_found":
            bucket["status"] = "not_found"
        if item.note:
            notes = bucket["notes"]
            assert isinstance(notes, list)
            if item.note not in notes:
                notes.append(item.note)

        details.append(
            [
                item.row_number,
                item.article,
                item.column,
                report_status_label(item.status),
                item.value,
                source_label(item.source),
                item.note,
            ]
        )
        if item.status in {"missing_value", "not_found"}:
            product_manager.append(
                [item.row_number, item.article, item.column, product_manager_action(item.column), item.note]
            )

    for (row_number, article), bucket in sorted(by_row.items()):
        notes = bucket["notes"]
        assert isinstance(notes, list)
        summary.append(
            [
                row_number,
                article,
                report_status_label(str(bucket["status"])),
                bucket["filled"],
                bucket["suggested"],
                bucket["missing"],
                "; ".join(notes),
            ]
        )

    for sheet in (summary, details, product_manager):
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 70)

    workbook.save(path)


def fill_template(
    db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None = None
) -> FillResult:
    suffix = template_path.suffix.lower()
    if suffix == ".csv":
        return fill_csv_template(db_path, template_path, output_dir, rules_path)
    if suffix in {".xlsx", ".xlsm"}:
        return fill_xlsx_template(db_path, template_path, output_dir, rules_path)
    raise ValueError("Поддерживаются только CSV, XLSX и XLSM шаблоны.")


def fill_csv_template(
    db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None = None
) -> FillResult:
    with template_path.open("r", newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except csv.Error:
            dialect = csv.excel()
            dialect.delimiter = ";"
        reader = list(csv.reader(fh, dialect))

    if not reader:
        raise ValueError("Шаблон пустой.")

    headers = [normalize_header(h) for h in reader[0]]
    article_idx = find_article_index(headers)
    if article_idx is None:
        raise ValueError("Не найдена колонка Артикул или Расширенный артикул.")

    articles = [normalize_article(row[article_idx]) for row in reader[1:] if len(row) > article_idx]
    products = fetch_products(db_path, articles)
    report: list[FillReportRow] = []
    filled_cells = 0
    suggested_cells = 0

    for row_number, row in enumerate(reader[1:], start=2):
        while len(row) < len(headers):
            row.append("")
        article = normalize_article(row[article_idx])
        if not article:
            continue
        product = products.get(article)
        if not product:
            report.append(FillReportRow(row_number, article, "", "not_found", "", "", "Артикул не найден в базе"))
            continue
        for col_idx, header in enumerate(headers):
            if header == "Артикул":
                continue
            value, status, source = product_value(product, header, rules_path)
            if status == "filled" and value:
                row[col_idx] = value
                filled_cells += 1
                report.append(FillReportRow(row_number, article, header, status, value, source))
            elif status == "suggested" and value:
                row[col_idx] = value
                filled_cells += 1
                suggested_cells += 1
                report.append(
                    FillReportRow(
                        row_number,
                        article,
                        header,
                        "filled_suggested",
                        value,
                        source,
                        "Заполнено по близкому совпадению; желательно проверить",
                    )
                )
            elif status in {"blank", "filled"} and is_reportable_header(header):
                report.append(
                    FillReportRow(
                        row_number,
                        article,
                        header,
                        "missing_value",
                        "",
                        source,
                        missing_note(header),
                    )
                )

    output_path, report_path = make_output_paths(template_path, output_dir)
    with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=getattr(dialect, "delimiter", ";"))
        writer.writerows(reader)
    write_report(report_path, report)

    unique_articles = {a for a in articles if a}
    found = len(unique_articles.intersection(products.keys()))
    return FillResult(
        output_path,
        report_path,
        len(reader) - 1,
        found,
        filled_cells,
        suggested_cells,
        len(unique_articles) - found,
    )


def fill_xlsx_template(
    db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None = None
) -> FillResult:
    workbook = load_workbook(template_path)
    sheet = workbook.active
    header_cells = list(sheet[1])
    headers = [normalize_header(cell.value) for cell in header_cells]
    fillable_columns = {
        idx
        for idx, cell in enumerate(header_cells)
        if is_yellow_header_cell(cell)
    }
    article_idx = find_article_index(headers)
    if article_idx is None:
        raise ValueError("Не найдена колонка Артикул или Расширенный артикул.")

    articles: list[str] = []
    for row in sheet.iter_rows(min_row=2):
        if article_idx < len(row):
            articles.append(normalize_article(row[article_idx].value))

    products = fetch_products(db_path, articles)
    report: list[FillReportRow] = []
    filled_cells = 0
    suggested_cells = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        article = normalize_article(row[article_idx].value if article_idx < len(row) else "")
        if not article:
            continue
        product = products.get(article)
        if not product:
            report.append(FillReportRow(row_number, article, "", "not_found", "", "", "Артикул не найден в базе"))
            for col_idx in fillable_columns:
                if col_idx != article_idx and col_idx < len(row):
                    mark_missing_cell(row[col_idx])
            continue
        for col_idx, header in enumerate(headers):
            if header == "Артикул" or col_idx >= len(row):
                continue
            if col_idx not in fillable_columns:
                continue
            value, status, source = product_value(product, header, rules_path)
            if status == "filled" and value:
                row[col_idx].value = value
                filled_cells += 1
                report.append(FillReportRow(row_number, article, header, status, value, source))
            elif status == "suggested" and value:
                row[col_idx].value = value
                filled_cells += 1
                suggested_cells += 1
                report.append(
                    FillReportRow(
                        row_number,
                        article,
                        header,
                        "filled_suggested",
                        value,
                        source,
                        "Заполнено по близкому совпадению; желательно проверить",
                    )
                )
            elif status in {"blank", "filled"} and is_reportable_header(header):
                mark_missing_cell(row[col_idx])
                report.append(
                    FillReportRow(
                        row_number,
                        article,
                        header,
                        "missing_value",
                        "",
                        source,
                        missing_note(header),
                    )
                )

    output_path, report_path = make_output_paths(template_path, output_dir)
    workbook.save(output_path)
    preserve_worksheet_dropdowns(template_path, output_path)
    write_report(report_path, report)

    unique_articles = {a for a in articles if a}
    found = len(unique_articles.intersection(products.keys()))
    return FillResult(
        output_path=output_path,
        report_path=report_path,
        total_rows=sheet.max_row - 1,
        found_articles=found,
        filled_cells=filled_cells,
        suggested_cells=suggested_cells,
        missing_articles=len(unique_articles) - found,
    )
