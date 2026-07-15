from __future__ import annotations

import datetime as dt
from collections import Counter, defaultdict
from dataclasses import dataclass, replace
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .db import ProductRecord, fetch_products
from .filler import (
    comparable_text,
    find_article_index,
    is_reportable_header,
    is_yellow_header_cell,
    missing_note,
    normalize_article,
    normalize_header,
    product_value,
    source_label,
)
from .mapping_rules import rejected_sources_for, rules_for, source_attributes_for


@dataclass(frozen=True)
class CandidateDetail:
    source: str
    count: int
    examples: tuple[str, ...]
    articles: tuple[str, ...]


@dataclass(frozen=True)
class FieldCoverage:
    class81_code: str
    field: str
    products_count: int
    exact_count: int
    approved_rule_count: int
    candidate_count: int
    filled_direct_count: int
    missing_count: int
    approved_sources: tuple[str, ...]
    candidate_sources: tuple[str, ...]
    candidate_details: tuple[CandidateDetail, ...]
    note: str


@dataclass(frozen=True)
class ArticleIssue:
    issue_type: str
    class81_code: str
    article: str
    field: str
    needed: str
    reason: str
    source: str = ""
    value: str = ""


def analyze_template_mapping(
    db_path: Path, template_path: Path, output_dir: Path, rules_path: Path | None = None
) -> Path:
    suffix = template_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Анализ маппинга сейчас поддерживает только XLSX/XLSM, потому что нужен цвет заголовков.")

    workbook = load_workbook(template_path, read_only=False, data_only=True)
    sheet = workbook.active
    header_cells = list(sheet[1])
    headers = [normalize_header(cell.value) for cell in header_cells]
    fillable_columns = {
        idx
        for idx, cell in enumerate(header_cells)
        if is_yellow_header_cell(cell)
    }
    article_idx = find_article_index(headers)
    if article_idx is None:
        raise ValueError("Не найдена колонка Артикул или Расширенный артикул.")
    class81_idx = headers.index("81 класс") if "81 класс" in headers else None

    articles: list[str] = []
    template_classes: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2):
        if article_idx < len(row):
            article = normalize_article(row[article_idx].value)
            articles.append(article)
            if article and class81_idx is not None and class81_idx < len(row):
                template_class = normalize_header(row[class81_idx].value)
                if template_class:
                    template_classes.setdefault(article, template_class)

    products = fetch_products(db_path, articles)
    by_class: dict[str, list[ProductRecord]] = defaultdict(list)
    for article in articles:
        if not article:
            continue
        product = products.get(article)
        if product:
            template_class = template_classes.get(article, "")
            if template_class and not product.class81_code:
                product = replace(product, class81_code=template_class)
            by_class[product.class81_code or ""].append(product)

    coverages: list[FieldCoverage] = []
    examples: list[list[str | int]] = []
    article_issues: list[ArticleIssue] = []

    yellow_headers = [
        header
        for idx, header in enumerate(headers)
        if idx in fillable_columns and header and header not in {"Артикул", "Расширенный артикул"}
    ]

    for class81_code, class_products in sorted(by_class.items()):
        for header in yellow_headers:
            if header.startswith("Конфиг:"):
                coverage, field_examples, field_issues = analyze_config_field(
                    class81_code, header, class_products, rules_path
                )
            elif is_reportable_header(header):
                coverage, field_examples, field_issues = analyze_direct_field(
                    class81_code, header, class_products, rules_path
                )
            else:
                coverage = FieldCoverage(
                    class81_code=class81_code,
                    field=header,
                    products_count=len(class_products),
                    exact_count=0,
                    approved_rule_count=0,
                    candidate_count=0,
                    filled_direct_count=0,
                    missing_count=len(class_products),
                    approved_sources=(),
                    candidate_sources=(),
                    candidate_details=(),
                    note="Поле пока не поддерживается логикой заполнения",
                )
                field_examples = []
                field_issues = []
            coverages.append(coverage)
            examples.extend(field_examples)
            article_issues.extend(field_issues)

    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_path = output_dir / f"{template_path.stem}_mapping_review_{stamp}.xlsx"
    write_mapping_report(report_path, coverages, examples, article_issues)
    return report_path


def analyze_config_field(
    class81_code: str,
    field: str,
    products: list[ProductRecord],
    rules_path: Path | None = None,
) -> tuple[FieldCoverage, list[list[str | int]], list[ArticleIssue]]:
    attr_name = field.removeprefix("Конфиг:").strip()
    approved_sources = source_attributes_for(class81_code, field, rules_path)
    approved_count = 0
    exact_count = 0
    candidate_count = 0
    missing_count = 0
    candidate_counter: Counter[str] = Counter()
    candidate_values: dict[str, list[str]] = defaultdict(list)
    candidate_articles: dict[str, list[str]] = defaultdict(list)
    example_rows: list[list[str | int]] = []
    article_issues: list[ArticleIssue] = []

    wanted = comparable_text(attr_name)
    rejected_sources = set(rejected_sources_for(class81_code, field, rules_path))
    for product in products:
        attrs = product.attributes or {}
        if attrs.get(attr_name):
            exact_count += 1
            example_rows.append([class81_code, field, attr_name, product.article, attrs[attr_name], "exact"])
            continue

        approved_value = ""
        approved_source = ""
        for source in approved_sources:
            if attrs.get(source):
                approved_value = attrs[source]
                approved_source = source
                break
        if approved_value:
            approved_count += 1
            example_rows.append(
                [class81_code, field, approved_source, product.article, approved_value, "approved_class_rule"]
            )
            continue

        candidates = [
            candidate
            for candidate in find_attribute_candidates(attr_name, attrs)
            if candidate[0] not in rejected_sources
        ]
        if candidates:
            candidate_count += 1
            source_name, value = candidates[0]
            candidate_counter[source_name] += 1
            if value not in candidate_values[source_name]:
                candidate_values[source_name].append(value)
            if product.article not in candidate_articles[source_name]:
                candidate_articles[source_name].append(product.article)
            example_rows.append([class81_code, field, source_name, product.article, value, "candidate"])
            article_issues.append(
                ArticleIssue(
                    issue_type="mapping",
                    class81_code=class81_code,
                    article=product.article,
                    field=field,
                    needed=f"Проверить источник для {field}",
                    reason="Есть похожее поле в базе, но нужно проверить, подходит ли оно по смыслу",
                    source=source_name,
                    value=value,
                )
            )
            continue

        missing_count += 1

    known_rules = rules_for(class81_code, field, rules_path)
    note = "Источник уже выбран" if known_rules else "Нужно выбрать источник вручную"
    if not wanted:
        note = "Пустое имя Конфиг-поля"

    candidate_details = tuple(
        CandidateDetail(
            source=name,
            count=count,
            examples=tuple(candidate_values[name][:5]),
            articles=tuple(candidate_articles[name][:5]),
        )
        for name, count in candidate_counter.most_common(8)
    )

    return (
        FieldCoverage(
            class81_code=class81_code,
            field=field,
            products_count=len(products),
            exact_count=exact_count,
            approved_rule_count=approved_count,
            candidate_count=candidate_count,
            filled_direct_count=0,
            missing_count=missing_count,
            approved_sources=approved_sources,
            candidate_sources=tuple(item.source for item in candidate_details),
            candidate_details=candidate_details,
            note=note,
        ),
        example_rows[:50],
        article_issues,
    )


def analyze_direct_field(
    class81_code: str,
    field: str,
    products: list[ProductRecord],
    rules_path: Path | None = None,
) -> tuple[FieldCoverage, list[list[str | int]], list[ArticleIssue]]:
    filled_count = 0
    missing_count = 0
    sources: Counter[str] = Counter()
    example_rows: list[list[str | int]] = []
    article_issues: list[ArticleIssue] = []

    for product in products:
        value, status, source = product_value(product, field, rules_path)
        if status == "filled" and value:
            filled_count += 1
            sources[source] += 1
            example_rows.append([class81_code, field, source, product.article, value, "filled"])
        else:
            missing_count += 1

    return (
        FieldCoverage(
            class81_code=class81_code,
            field=field,
            products_count=len(products),
            exact_count=0,
            approved_rule_count=0,
            candidate_count=0,
            filled_direct_count=filled_count,
            missing_count=missing_count,
            approved_sources=tuple(name for name, _ in sources.most_common(5)),
            candidate_sources=(),
            candidate_details=(),
            note="Обычное поле из базы",
        ),
        example_rows[:50],
        article_issues,
    )


def find_attribute_candidates(attr_name: str, attrs: dict[str, str]) -> list[tuple[str, str]]:
    wanted = comparable_text(attr_name)
    if not wanted:
        return []

    candidates: list[tuple[int, str, str]] = []
    for source_name, value in attrs.items():
        source_key = comparable_text(source_name)
        if not source_key or not value:
            continue
        if wanted == source_key:
            score = 100
        elif wanted in source_key or source_key in wanted:
            score = 70
        else:
            wanted_tokens = set(attr_name.lower().replace(",", " ").split())
            source_tokens = set(source_name.lower().replace(",", " ").split())
            score = len(wanted_tokens.intersection(source_tokens)) * 10
        if score > 0:
            candidates.append((score, source_name, value))

    candidates.sort(key=lambda item: (-item[0], item[1]))
    return [(source_name, value) for _, source_name, value in candidates[:5]]


def coverage_status(item: FieldCoverage) -> tuple[str, int, int, int]:
    will_fill = item.filled_direct_count + item.exact_count + item.approved_rule_count
    needs_mapping = item.candidate_count
    missing_count = item.missing_count
    if item.products_count == 0:
        status = "Нет товаров"
    elif will_fill == item.products_count:
        status = "Заполнится"
    elif will_fill == 0 and needs_mapping and not missing_count:
        status = "Нужен выбор источника"
    elif will_fill == 0 and missing_count == item.products_count:
        status = "Нет данных в базе"
    elif will_fill > 0 and missing_count == 0 and needs_mapping:
        status = "Частично: нужен выбор источника"
    elif will_fill > 0 and missing_count > 0:
        status = "Частично заполнится"
    else:
        status = "Смешанный статус"
    return status, will_fill, needs_mapping, missing_count


def example_status_label(status: str) -> str:
    labels = {
        "filled": "Заполнится",
        "exact": "Заполнится",
        "approved_class_rule": "Заполнится",
        "candidate": "Нужно проверить источник",
    }
    return labels.get(status, status)


def review_source_label(source: str) -> str:
    if not source:
        return ""
    if "." in source or "/" in source or source in {"static", "generated"}:
        return source_label(source)
    return source


def write_mapping_report(
    path: Path,
    coverages: list[FieldCoverage],
    examples: list[list[str | int]],
    article_issues: list[ArticleIssue],
) -> None:
    workbook = Workbook()
    coverage_sheet = workbook.active
    coverage_sheet.title = "Покрытие"
    coverage_sheet.append(
        [
            "Категория",
            "Поле шаблона",
            "Статус",
            "Товаров",
            "Заполнится",
            "Нужен выбор источника",
            "Не заполнится",
            "Комментарий",
        ]
    )
    for item in coverages:
        status, will_fill, needs_mapping, missing_count = coverage_status(item)
        coverage_sheet.append(
            [
                item.class81_code,
                item.field,
                status,
                item.products_count,
                will_fill,
                needs_mapping,
                missing_count,
                item.note,
            ]
        )

    examples_sheet = workbook.create_sheet("Примеры")
    examples_sheet.append(["Категория", "Поле шаблона", "Источник", "Артикул", "Значение", "Состояние"])
    for row in examples:
        display_row = list(row)
        display_row[2] = review_source_label(str(display_row[2] or ""))
        display_row[5] = example_status_label(str(display_row[5] or ""))
        examples_sheet.append(display_row)

    mapping_sheet = workbook.create_sheet("Выбор источника")
    mapping_sheet.append(
        [
            "Категория",
            "Артикул",
            "Поле шаблона",
            "Возможный источник",
            "Значение",
            "Комментарий",
        ]
    )
    for item in article_issues:
        if item.issue_type != "mapping":
            continue
        mapping_sheet.append(
            [item.class81_code, item.article, item.field, item.source, item.value, item.reason]
        )

    rules_sheet = workbook.create_sheet("Правила")
    rules_sheet.append(
        [
            "81 класс",
            "Поле шаблона",
            "Кандидат источника",
            "Покрытие кандидата",
            "Примеры значений",
            "Примеры артикулов",
            "Действие",
        ]
    )
    for item in coverages:
        if not item.field.startswith("Конфиг:"):
            continue
        if item.approved_sources:
            rules_sheet.append(
                [
                    item.class81_code,
                    item.field,
                    "; ".join(item.approved_sources),
                    item.approved_rule_count,
                    "",
                    "",
                    "уже утверждено",
                ]
            )
            continue
        for candidate in item.candidate_details:
            rules_sheet.append(
                [
                    item.class81_code,
                    item.field,
                    candidate.source,
                    candidate.count,
                    "; ".join(candidate.examples),
                    "; ".join(candidate.articles),
                    "проверить и принять, если источник подходит",
                ]
            )

    for sheet in (coverage_sheet, examples_sheet, mapping_sheet, rules_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 80)

    workbook.save(path)
