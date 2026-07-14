from __future__ import annotations

import datetime as dt
import re
import shutil
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .price_importer import backup_database


ARTICLE_HINT = "артикул"
DIMENSION_HINTS = {
    "length_mm": ("глубин", "длин"),
    "width_mm": ("ширин",),
    "height_mm": ("высот",),
}
DIMENSION_EXCLUDE_HINTS = {
    "length_mm": (),
    "width_mm": (),
    "height_mm": ("уровнем моря", "над уровнем", "altitude"),
}
OVERALL_DIMENSION_HINT = "габарит"
OVERALL_DIMENSION_SECONDARY_HINT = "размер"
SOURCE_KIND = "etim_manual"


@dataclass(frozen=True)
class EtimArticleData:
    article: str
    sheet_name: str
    row_number: int
    dimensions: dict[str, float]
    dimension_sources: dict[str, str]
    attributes: dict[str, str]


@dataclass(frozen=True)
class EtimImportResult:
    etim_file: Path
    db_path: Path
    backup_path: Path | None
    import_id: int
    scanned_articles: int
    found_articles: int
    dimension_values_found: int
    dimension_values_written: int
    dimension_conflicts: int
    attributes_written: int


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return " ".join(text.split())


def extract_first_number(value: object) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return None


def parse_overall_dimensions(value: object) -> dict[str, float]:
    text = normalize_text(value)
    if not text:
        return {}

    chains = re.findall(
        r"(\d+(?:[.,]\d+)?(?:\s*[xх×*]\s*\d+(?:[.,]\d+)?){1,})",
        text,
        flags=re.IGNORECASE,
    )
    if not chains:
        return {}

    best_chain = max(chains, key=lambda chain: len(re.findall(r"\d+(?:[.,]\d+)?", chain)))
    values_mm = [
        float(value.replace(",", "."))
        for value in re.findall(r"\d+(?:[.,]\d+)?", best_chain)
    ]
    if len(values_mm) >= 4:
        values_mm = values_mm[-3:]
    if len(values_mm) < 2:
        return {}

    dimensions = {
        "width_mm": values_mm[0],
        "height_mm": values_mm[1],
    }
    if len(values_mm) >= 3:
        dimensions["length_mm"] = values_mm[2]
    return dimensions


def parse_dimension_cell(value: object, header: object, dimension_key: str) -> float | None:
    number = extract_first_number(value)
    if number is None:
        return None

    header_text = normalize_text(header).casefold()
    if any(hint in header_text for hint in DIMENSION_EXCLUDE_HINTS.get(dimension_key, ())):
        return None
    if "мм" in header_text or "mm" in header_text:
        return number
    if re.search(r"\bм\b", header_text):
        return number * 1000.0
    return number


def find_dimension_column(headers: list[object], hints: Iterable[str], exclude_hints: Iterable[str]) -> int | None:
    normalized = [normalize_text(header).casefold() for header in headers]
    for index, header in enumerate(normalized):
        if (
            any(hint in header for hint in hints)
            and not any(exclude in header for exclude in exclude_hints)
            and ("мм" in header or "mm" in header)
        ):
            return index
    for index, header in enumerate(normalized):
        if (
            any(hint in header for hint in hints)
            and not any(exclude in header for exclude in exclude_hints)
            and any(marker in header for marker in ("габарит", "размер", "dimension"))
        ):
            return index
    return None


def find_overall_dimension_column(headers: list[object]) -> int | None:
    normalized = [normalize_text(header).casefold() for header in headers]
    for index, header in enumerate(normalized):
        if OVERALL_DIMENSION_HINT in header and OVERALL_DIMENSION_SECONDARY_HINT in header:
            return index
    for index, header in enumerate(normalized):
        if OVERALL_DIMENSION_HINT in header:
            return index
    return None


def find_header_row(rows: list[tuple[object, ...]], row_index: int, column_index: int) -> int | None:
    for index in range(row_index - 1, -1, -1):
        if column_index < len(rows[index]) and ARTICLE_HINT in normalize_text(rows[index][column_index]).casefold():
            return index
    return None


def collect_attributes(headers: list[object], row: tuple[object, ...], article_column: int) -> dict[str, str]:
    attrs: dict[str, str] = {}
    for index, header_value in enumerate(headers):
        if index == article_column:
            continue
        header = normalize_text(header_value)
        if not header or ARTICLE_HINT == header.casefold():
            continue
        value = normalize_text(row[index] if index < len(row) else "")
        if value:
            attrs.setdefault(header, value)
    return attrs


def scan_etim_workbook(etim_path: Path, articles: Iterable[str]) -> dict[str, EtimArticleData]:
    if etim_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("ETIM-файл должен быть XLSX/XLSM.")

    articles_to_find = {normalize_text(article) for article in articles if normalize_text(article)}
    if not articles_to_find:
        return {}

    workbook = load_workbook(etim_path, read_only=True, data_only=True)
    found: dict[str, EtimArticleData] = {}
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            for row_index, row in enumerate(rows):
                if len(found) == len(articles_to_find):
                    return found
                for column_index, cell_value in enumerate(row):
                    article = normalize_text(cell_value)
                    if article not in articles_to_find or article in found:
                        continue
                    header_row_index = find_header_row(rows, row_index, column_index)
                    if header_row_index is None:
                        continue
                    headers = list(rows[header_row_index])
                    dimensions: dict[str, float] = {}
                    dimension_sources: dict[str, str] = {}

                    overall_column = find_overall_dimension_column(headers)
                    if overall_column is not None and overall_column < len(row):
                        parsed = parse_overall_dimensions(row[overall_column])
                        for key, value in parsed.items():
                            dimensions[key] = value
                            dimension_sources[key] = normalize_text(headers[overall_column])

                    for key, hints in DIMENSION_HINTS.items():
                        if key in dimensions:
                            continue
                        column = find_dimension_column(headers, hints, DIMENSION_EXCLUDE_HINTS.get(key, ()))
                        if column is None or column >= len(row):
                            continue
                        parsed_value = parse_dimension_cell(row[column], headers[column], key)
                        if parsed_value is not None:
                            dimensions[key] = parsed_value
                            dimension_sources[key] = normalize_text(headers[column])

                    found[article] = EtimArticleData(
                        article=article,
                        sheet_name=sheet.title,
                        row_number=row_index + 1,
                        dimensions=dimensions,
                        dimension_sources=dimension_sources,
                        attributes=collect_attributes(headers, row, column_index),
                    )
    finally:
        workbook.close()
    return found


def ensure_etim_import_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dimension_imports (
            id INTEGER PRIMARY KEY,
            source_kind TEXT NOT NULL,
            local_file TEXT,
            workbook_name TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_dimensions_raw (
            id INTEGER PRIMARY KEY,
            import_id INTEGER,
            article TEXT NOT NULL,
            source_kind TEXT NOT NULL,
            dimension_kind TEXT NOT NULL,
            length_mm REAL,
            width_mm REAL,
            height_mm REAL,
            weight_kg REAL,
            volume_m3 REAL,
            source_sheet TEXT,
            source_field TEXT,
            value_raw TEXT,
            confidence REAL DEFAULT 1.0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(import_id) REFERENCES dimension_imports(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_dimensions_resolved (
            article TEXT PRIMARY KEY,
            length_mm REAL,
            width_mm REAL,
            height_mm REAL,
            weight_kg REAL,
            volume_m3 REAL,
            dimensions_source TEXT,
            weight_source TEXT,
            volume_source TEXT,
            last_resolved_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_attribute_values (
            id INTEGER PRIMARY KEY,
            article TEXT NOT NULL,
            attribute_code TEXT,
            attribute_name TEXT,
            sheet_name TEXT,
            etim_class_code TEXT,
            value_raw TEXT,
            value_normalized TEXT,
            attribute_group TEXT,
            source_kind TEXT,
            source_ref TEXT,
            confidence REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


def fetch_all_articles(conn: sqlite3.Connection) -> list[str]:
    return [
        str(row[0])
        for row in conn.execute(
            "SELECT article FROM products WHERE COALESCE(article, '') <> '' ORDER BY article"
        ).fetchall()
    ]


def count_written_and_conflicts(
    conn: sqlite3.Connection, article_data: dict[str, EtimArticleData]
) -> tuple[int, int]:
    written = 0
    conflicts = 0
    for data in article_data.values():
        if not data.dimensions:
            continue
        row = conn.execute(
            """
            SELECT length_mm, width_mm, height_mm
            FROM product_dimensions_resolved
            WHERE article = ?
            """,
            (data.article,),
        ).fetchone()
        existing = dict(row) if row else {}
        for key, value in data.dimensions.items():
            current = existing.get(key)
            if current is None:
                written += 1
            elif abs(float(current) - float(value)) > 0.001:
                conflicts += 1
    return written, conflicts


def insert_etim_data(
    conn: sqlite3.Connection,
    import_id: int,
    article_data: dict[str, EtimArticleData],
    source_ref: str,
) -> tuple[int, int]:
    dimension_values_found = 0
    attributes_written = 0
    for data in article_data.values():
        for key, value in data.dimensions.items():
            dimension_values_found += 1
            length_mm = value if key == "length_mm" else None
            width_mm = value if key == "width_mm" else None
            height_mm = value if key == "height_mm" else None
            source_field = data.dimension_sources.get(key, "")
            conn.execute(
                """
                INSERT INTO product_dimensions_raw (
                    import_id, article, source_kind, dimension_kind,
                    length_mm, width_mm, height_mm, source_sheet, source_field,
                    value_raw, confidence
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    data.article,
                    SOURCE_KIND,
                    key,
                    length_mm,
                    width_mm,
                    height_mm,
                    data.sheet_name,
                    source_field,
                    str(value),
                    0.95,
                ),
            )
        if data.dimensions:
            conn.execute(
                """
                INSERT INTO product_dimensions_resolved (
                    article, length_mm, width_mm, height_mm, dimensions_source, last_resolved_at
                ) VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(article) DO UPDATE SET
                    length_mm=COALESCE(product_dimensions_resolved.length_mm, excluded.length_mm),
                    width_mm=COALESCE(product_dimensions_resolved.width_mm, excluded.width_mm),
                    height_mm=COALESCE(product_dimensions_resolved.height_mm, excluded.height_mm),
                    dimensions_source=CASE
                        WHEN product_dimensions_resolved.dimensions_source IS NULL
                             OR product_dimensions_resolved.dimensions_source = ''
                        THEN excluded.dimensions_source
                        ELSE product_dimensions_resolved.dimensions_source
                    END,
                    last_resolved_at=CURRENT_TIMESTAMP
                """,
                (
                    data.article,
                    data.dimensions.get("length_mm"),
                    data.dimensions.get("width_mm"),
                    data.dimensions.get("height_mm"),
                    SOURCE_KIND,
                ),
            )

        for attribute_name, value in data.attributes.items():
            conn.execute(
                """
                INSERT INTO product_attribute_values (
                    article, attribute_name, sheet_name, value_raw, value_normalized,
                    source_kind, source_ref, confidence
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data.article,
                    attribute_name,
                    data.sheet_name,
                    value,
                    value,
                    SOURCE_KIND,
                    source_ref,
                    0.8,
                ),
            )
            attributes_written += 1
    return dimension_values_found, attributes_written


def import_etim_workbook(etim_path: Path, db_path: Path, make_backup: bool = True) -> EtimImportResult:
    if not etim_path.exists():
        raise FileNotFoundError(f"ETIM-файл не найден: {etim_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"База не найдена: {db_path}")

    backup_path = backup_database(db_path) if make_backup else None
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        ensure_etim_import_tables(conn)
        articles = fetch_all_articles(conn)
        scanned_articles = len(articles)

    article_data = scan_etim_workbook(etim_path, articles)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        ensure_etim_import_tables(conn)
        dimension_values_written, dimension_conflicts = count_written_and_conflicts(conn, article_data)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO dimension_imports (source_kind, local_file, workbook_name, notes)
            VALUES (?, ?, ?, ?)
            """,
            (
                SOURCE_KIND,
                str(etim_path),
                etim_path.name,
                f"ETIM import {dt.datetime.now().isoformat(timespec='seconds')}",
            ),
        )
        import_id = int(cur.lastrowid)
        dimension_values_found, attributes_written = insert_etim_data(
            conn, import_id, article_data, etim_path.name
        )
        conn.commit()

    return EtimImportResult(
        etim_file=etim_path,
        db_path=db_path,
        backup_path=backup_path,
        import_id=import_id,
        scanned_articles=scanned_articles,
        found_articles=len(article_data),
        dimension_values_found=dimension_values_found,
        dimension_values_written=dimension_values_written,
        dimension_conflicts=dimension_conflicts,
        attributes_written=attributes_written,
    )


def check_sqlite_integrity(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        result = str(conn.execute("PRAGMA integrity_check").fetchone()[0])
    if result.lower() != "ok":
        raise ValueError(f"SQLite integrity_check failed for {db_path}: {result}")


def restore_database_backup(db_path: Path, backup_path: Path) -> Path:
    if not db_path.exists():
        raise FileNotFoundError(f"Текущая база не найдена: {db_path}")
    if not backup_path.exists():
        raise FileNotFoundError(f"Бэкап не найден: {backup_path}")

    check_sqlite_integrity(backup_path)
    safety_backup = db_path.with_name(
        f"{db_path.name}.before_restore_{dt.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    )
    shutil.copy2(db_path, safety_backup)
    shutil.copy2(backup_path, db_path)
    check_sqlite_integrity(db_path)
    return safety_backup


def result_lines(result: EtimImportResult) -> list[str]:
    lines = [
        f"ETIM-файл: {result.etim_file}",
        f"База: {result.db_path}",
        f"Импорт id: {result.import_id}",
        f"Артикулов в базе для поиска: {result.scanned_articles}",
        f"Артикулов найдено в ETIM: {result.found_articles}",
        f"Габаритных значений найдено: {result.dimension_values_found}",
        f"Новых габаритных значений записано: {result.dimension_values_written}",
        f"Конфликтов с уже заполненными габаритами: {result.dimension_conflicts}",
        f"Характеристик записано: {result.attributes_written}",
    ]
    if result.backup_path:
        lines.append(f"Бэкап базы: {result.backup_path}")
    return lines
